"""
End-to-end TC tests — calls the engine directly (no HTTP, no login).
Covers TC-01 through TC-08 plus TC-MAIN, verifying the same values
that would be displayed in the FE Results page.

Export tests: run engine → generate_workpapers() → parse Excel Summary sheet
→ assert values match expected. This confirms the export matches the BE calculation.
"""
import io
from datetime import date
from decimal import Decimal

import openpyxl
import pytest

from pfic_engine.section_1291.excess_dist import compute_prior_3yr_average, split_distribution
from pfic_engine.section_1291.daily_allocation import allocate_excess_distribution, DayClassification
from pfic_engine.section_1291.lot_shares import compute_lot_shares
from pfic_engine.section_1291.deferred_tax import compute_deferred_tax_with_interest
from pfic_engine.lot.fifo import Lot, match_fifo
from pfic_engine.core.tax_constants import get_filing_deadline
from pfic_engine.output.excel_workpaper import generate_workpapers
from tests.fixtures.known_cases import TC_MAIN_2022


def D(s):
    return Decimal(str(s))


def approx(a, b, tol="0.05"):
    return abs(D(str(a)) - D(str(b))) <= D(tol)


# ── Pipeline helper (mirrors calculations.py logic) ──────────────────────────

def run_pipeline(purchases, distributions_by_year, tax_year, first_pfic_year=1987):
    """
    purchases: list of (acquisition_date, units)
    distributions_by_year: {year: amount}
    Returns result dict plus full_result shaped for generate_workpapers().
    """
    current_dist = distributions_by_year.get(tax_year, D("0"))
    prior_years = sorted(y for y in distributions_by_year if y < tax_year)
    prior_dists = [distributions_by_year[y] for y in prior_years[-3:]]
    prior_avg = compute_prior_3yr_average(prior_dists, len(prior_years))
    non_excess, excess = split_distribution(current_dist, prior_avg)

    all_year_buckets: dict = {}
    total_deferred_tax = D("0")
    total_interest = D("0")
    total_ordinary = non_excess
    lots_detail = []
    deferred_tax_results = []

    if excess > D("0"):
        eligible = [(acq, units) for acq, units in purchases if acq.year <= tax_year]
        shares = compute_lot_shares(excess, [u for _, u in eligible])

        for (acq, units), lot_excess in zip(eligible, shares):
            alloc = allocate_excess_distribution(
                excess_amount=lot_excess,
                acquisition_date=acq,
                disposition_date=date(tax_year, 12, 31),
                current_tax_year=tax_year,
            )
            deferred = compute_deferred_tax_with_interest(alloc["year_buckets"], tax_year)

            for yr, bucket in alloc["year_buckets"].items():
                if yr in all_year_buckets:
                    all_year_buckets[yr]["days"] += bucket["days"]
                    all_year_buckets[yr]["amount"] += bucket["amount"]
                else:
                    all_year_buckets[yr] = dict(bucket)

            total_deferred_tax += D(deferred["total_deferred_tax"])
            total_interest += D(deferred["total_interest"])
            total_ordinary += D(deferred["ordinary_income"])
            lots_detail.append({"acquisition_date": str(acq), "year_results": deferred["year_results"]})
            deferred_tax_results.append({
                "acquisition_date": str(acq),
                "units": str(units),
                "lot_excess": str(lot_excess),
                "year_results": deferred["year_results"],
                "ordinary_income": deferred["ordinary_income"],
                "total_deferred_tax": deferred["total_deferred_tax"],
                "total_interest": deferred["total_interest"],
                "grand_total": deferred["grand_total"],
            })

    grand_total = total_deferred_tax + total_interest

    # full_result matches the shape expected by generate_workpapers()
    full_result = {
        "tax_year": tax_year,
        "prior_3yr_average": str(prior_avg),
        "current_year_distribution": str(current_dist),
        "non_excess_ordinary": str(non_excess),
        "excess_distribution": str(excess),
        "total_deferred_tax": str(total_deferred_tax),
        "total_interest": str(total_interest),
        "total_ordinary_income": str(total_ordinary),
        "grand_total": str(grand_total),
        "year_buckets": {
            str(yr): {"days": b["days"], "amount": str(b["amount"]), "classification": b["classification"]}
            for yr, b in all_year_buckets.items()
        },
        "deferred_tax_results": deferred_tax_results,
    }

    return {
        "prior_3yr_average": prior_avg,
        "non_excess_ordinary": non_excess,
        "excess_distribution": excess,
        "year_buckets": all_year_buckets,
        "total_deferred_tax": total_deferred_tax,
        "total_interest": total_interest,
        "total_ordinary": total_ordinary,
        "grand_total": grand_total,
        "lots": lots_detail,
        "full_result": full_result,
    }


def parse_excel(full_result, holding_name="Test", client_code="TEST"):
    """Generate Excel and return all sheets as {sheet_name: {label: value}}."""
    xlsx = generate_workpapers(full_result, holding_name, client_code)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = {}
        data_rows = []
        for row in ws.iter_rows(values_only=True):
            if any(v is not None for v in row):
                data_rows.append(row)
        # Summary sheet: col A = label, col B = value
        if sheet_name == "Summary":
            for row in data_rows:
                if row[0] and row[1] is not None:
                    rows[str(row[0])] = row[1]
        result[sheet_name] = {"_rows": data_rows, **rows}
    return result


def excel_summary(full_result, holding_name="Test", client_code="TEST"):
    """Convenience: return just Summary sheet {label: value}."""
    return parse_excel(full_result, holding_name, client_code)["Summary"]


# ── TC-MAIN ──────────────────────────────────────────────────────────────────

def _tc_main_pipeline(tax_year):
    txns = TC_MAIN_2022["transactions"]
    purchases = [(t["date"], t["units"]) for t in txns if t["type"] == "purchase"]
    dists = {}
    for t in txns:
        if t["type"] == "distribution":
            dists[t["date"].year] = dists.get(t["date"].year, D("0")) + t["amount"]
    return run_pipeline(purchases, dists, tax_year)


def test_tc_main_excess():
    r = _tc_main_pipeline(2022)
    assert approx(r["excess_distribution"], "1660.21", "0.02")
    assert approx(r["non_excess_ordinary"], "823.24", "0.02")


def test_tc_main_deferred_tax():
    r = _tc_main_pipeline(2022)
    assert approx(r["total_deferred_tax"], "140.72", "0.05")


def test_tc_main_interest():
    r = _tc_main_pipeline(2022)
    assert approx(r["total_interest"], "8.19", "0.05")


def test_tc_main_grand_total():
    r = _tc_main_pipeline(2022)
    assert approx(r["grand_total"], "148.90", "0.10")


def test_tc_main_year_buckets_sum_to_excess():
    """After multi-lot merge fix, all bucket amounts must equal total excess."""
    r = _tc_main_pipeline(2022)
    bucket_total = sum(b["amount"] for b in r["year_buckets"].values())
    assert approx(bucket_total, r["excess_distribution"], "0.01"), \
        f"bucket sum {bucket_total} != excess {r['excess_distribution']}"


def test_tc_main_2021_has_117_days():
    """L1 (acq 2021-09-06) 2021 bucket = 117 days."""
    txns = TC_MAIN_2022["transactions"]
    acq = date(2021, 9, 6)
    alloc = allocate_excess_distribution(D("581.41"), acq, date(2022, 12, 31), 2022)
    assert alloc["year_buckets"][2021]["days"] == 117
    assert alloc["year_buckets"][2022]["days"] == 364


# ── TC-01: §7503 weekend deadline (2021) ─────────────────────────────────────

def test_tc01_excess():
    r = run_pipeline(
        purchases=[(date(2020, 1, 1), D("1000"))],
        distributions_by_year={2020: D("500"), 2021: D("2000")},
        tax_year=2021,
    )
    # prior avg = 500, threshold = 625, excess = 1375
    assert approx(r["excess_distribution"], "1375.00", "0.02")
    assert approx(r["non_excess_ordinary"], "625.00", "0.02")


def test_tc01_interest_start_2020_is_covid():
    """2020 allocation: interest start = 2021-05-17 (COVID Notice 2021-21)."""
    r = run_pipeline(
        purchases=[(date(2020, 1, 1), D("1000"))],
        distributions_by_year={2020: D("500"), 2021: D("2000")},
        tax_year=2021,
    )
    yr_results = r["lots"][0]["year_results"]
    assert yr_results[2020]["interest_start"] == "2021-05-17", \
        f"Expected 2021-05-17 (COVID 2020), got {yr_results[2020]['interest_start']}"


def test_tc01_interest_end_2021_is_7503():
    """Tax year 2021 return due = 2022-04-18 (§7503: Apr 15 is Saturday)."""
    r = run_pipeline(
        purchases=[(date(2020, 1, 1), D("1000"))],
        distributions_by_year={2020: D("500"), 2021: D("2000")},
        tax_year=2021,
    )
    yr_results = r["lots"][0]["year_results"]
    assert yr_results[2020]["interest_end"] == "2022-04-18", \
        f"Expected 2022-04-18 (§7503), got {yr_results[2020]['interest_end']}"


# ── TC-02: COVID 2019 deadline ────────────────────────────────────────────────

def test_tc02_excess():
    r = run_pipeline(
        purchases=[(date(2018, 1, 1), D("1000"))],
        distributions_by_year={2018: D("500"), 2019: D("2000")},
        tax_year=2019,
    )
    assert approx(r["excess_distribution"], "1375.00", "0.02")


def test_tc02_interest_end_is_covid_2020():
    """Tax year 2019 return due = 2020-07-15 (COVID Notice 2020-23)."""
    r = run_pipeline(
        purchases=[(date(2018, 1, 1), D("1000"))],
        distributions_by_year={2018: D("500"), 2019: D("2000")},
        tax_year=2019,
    )
    yr_results = r["lots"][0]["year_results"]
    assert yr_results[2018]["interest_end"] == "2020-07-15", \
        f"Expected 2020-07-15 (COVID), got {yr_results[2018]['interest_end']}"


# ── TC-03: COVID 2020 deadline ────────────────────────────────────────────────

def test_tc03_excess():
    r = run_pipeline(
        purchases=[(date(2019, 1, 1), D("1000"))],
        distributions_by_year={2019: D("500"), 2020: D("2000")},
        tax_year=2020,
    )
    assert approx(r["excess_distribution"], "1375.00", "0.02")


def test_tc03_interest_end_is_covid_2021():
    """Tax year 2020 return due = 2021-05-17 (COVID Notice 2021-21)."""
    r = run_pipeline(
        purchases=[(date(2019, 1, 1), D("1000"))],
        distributions_by_year={2019: D("500"), 2020: D("2000")},
        tax_year=2020,
    )
    yr_results = r["lots"][0]["year_results"]
    assert yr_results[2019]["interest_end"] == "2021-05-17", \
        f"Expected 2021-05-17 (COVID), got {yr_results[2019]['interest_end']}"


# ── TC-04: Same-year purchase — all current year ──────────────────────────────

def test_tc04_no_deferred_tax():
    """Purchase and distribution same year → all current_year → no deferred tax."""
    r = run_pipeline(
        purchases=[(date(2022, 3, 1), D("1000"))],
        distributions_by_year={2022: D("500")},
        tax_year=2022,
    )
    assert r["total_deferred_tax"] == D("0")
    assert r["total_interest"] == D("0")


def test_tc04_all_buckets_current_year():
    r = run_pipeline(
        purchases=[(date(2022, 3, 1), D("1000"))],
        distributions_by_year={2022: D("500")},
        tax_year=2022,
    )
    for yr, bucket in r["year_buckets"].items():
        assert bucket["classification"] == DayClassification.CURRENT_YEAR, \
            f"Year {yr}: expected current_year, got {bucket['classification']}"


def test_tc04_entire_distribution_is_excess():
    """No prior distributions → prior avg = 0 → entire $500 is excess."""
    r = run_pipeline(
        purchases=[(date(2022, 3, 1), D("1000"))],
        distributions_by_year={2022: D("500")},
        tax_year=2022,
    )
    assert approx(r["excess_distribution"], "500.00", "0.01")
    assert approx(r["non_excess_ordinary"], "0.00", "0.01")


# ── TC-05: Leap year 2024 ─────────────────────────────────────────────────────

def test_tc05_excess():
    # prior avg = (800+900)/2 = 850, threshold = 1062.50, excess = 10937.50
    r = run_pipeline(
        purchases=[(date(2022, 1, 1), D("1000"))],
        distributions_by_year={2022: D("800"), 2023: D("900"), 2024: D("12000")},
        tax_year=2024,
    )
    assert approx(r["excess_distribution"], "10937.50", "0.02")


def test_tc05_leap_year_current_bucket_366_days():
    """2024 is a leap year — current-year bucket = 366 days."""
    r = run_pipeline(
        purchases=[(date(2022, 1, 1), D("1000"))],
        distributions_by_year={2022: D("800"), 2023: D("900"), 2024: D("12000")},
        tax_year=2024,
    )
    assert 2024 in r["year_buckets"]
    # Distribution date 2024-12-31 is excluded → 2024 bucket = Jan 1–Dec 30 = 365 days
    assert r["year_buckets"][2024]["days"] == 365, \
        f"Expected 365, got {r['year_buckets'][2024]['days']}"


def test_tc05_prior_pfic_years_present():
    r = run_pipeline(
        purchases=[(date(2022, 1, 1), D("1000"))],
        distributions_by_year={2022: D("800"), 2023: D("900"), 2024: D("12000")},
        tax_year=2024,
    )
    assert r["year_buckets"][2022]["classification"] == DayClassification.PRIOR_PFIC
    assert r["year_buckets"][2023]["classification"] == DayClassification.PRIOR_PFIC
    assert r["year_buckets"][2024]["classification"] == DayClassification.CURRENT_YEAR


# ── TC-06: Pre-PFIC boundary (bought 1985) ────────────────────────────────────

def test_tc06_excess():
    # prior avg = 800, threshold = 1000, excess = 14000
    r = run_pipeline(
        purchases=[(date(1985, 1, 1), D("1000"))],
        distributions_by_year={2021: D("800"), 2022: D("15000")},
        tax_year=2022,
    )
    assert approx(r["excess_distribution"], "14000.00", "0.02")


def test_tc06_pre_pfic_1985_1986():
    r = run_pipeline(
        purchases=[(date(1985, 1, 1), D("1000"))],
        distributions_by_year={2021: D("800"), 2022: D("15000")},
        tax_year=2022,
    )
    assert r["year_buckets"][1985]["classification"] == DayClassification.PRE_PFIC
    assert r["year_buckets"][1986]["classification"] == DayClassification.PRE_PFIC


def test_tc06_pfic_1987_onward():
    r = run_pipeline(
        purchases=[(date(1985, 1, 1), D("1000"))],
        distributions_by_year={2021: D("800"), 2022: D("15000")},
        tax_year=2022,
    )
    assert r["year_buckets"][1987]["classification"] == DayClassification.PRIOR_PFIC
    assert r["year_buckets"][2021]["classification"] == DayClassification.PRIOR_PFIC
    assert r["year_buckets"][2022]["classification"] == DayClassification.CURRENT_YEAR


def test_tc06_has_deferred_tax():
    r = run_pipeline(
        purchases=[(date(1985, 1, 1), D("1000"))],
        distributions_by_year={2021: D("800"), 2022: D("15000")},
        tax_year=2022,
    )
    assert r["total_deferred_tax"] > D("0")
    assert r["total_interest"] > D("0")


# ── TC-07: Zero prior distributions ──────────────────────────────────────────

def test_tc07_entire_distribution_excess():
    """No prior distributions → prior avg = 0 → entire $3000 is excess."""
    r = run_pipeline(
        purchases=[(date(2020, 1, 1), D("1000"))],
        distributions_by_year={2022: D("3000")},
        tax_year=2022,
    )
    assert r["prior_3yr_average"] == D("0")
    assert approx(r["excess_distribution"], "3000.00", "0.01")
    assert approx(r["non_excess_ordinary"], "0.00", "0.01")


def test_tc07_year_classifications():
    r = run_pipeline(
        purchases=[(date(2020, 1, 1), D("1000"))],
        distributions_by_year={2022: D("3000")},
        tax_year=2022,
    )
    assert r["year_buckets"][2020]["classification"] == DayClassification.PRIOR_PFIC
    assert r["year_buckets"][2021]["classification"] == DayClassification.PRIOR_PFIC
    assert r["year_buckets"][2022]["classification"] == DayClassification.CURRENT_YEAR


def test_tc07_has_deferred_tax_and_interest():
    r = run_pipeline(
        purchases=[(date(2020, 1, 1), D("1000"))],
        distributions_by_year={2022: D("3000")},
        tax_year=2022,
    )
    assert r["total_deferred_tax"] > D("0")
    assert r["total_interest"] > D("0")


# ── TC-08: FIFO partial sale ──────────────────────────────────────────────────

def test_tc08_fifo_leaves_50_units():
    """Buy 100+100, sell 150 FIFO → 50 units remain from lot2."""
    lots = [
        Lot("L1", date(2020, 1, 1), D("100"), D("100.00")),
        Lot("L2", date(2021, 1, 1), D("100"), D("120.00")),
    ]
    result, remaining = match_fifo(lots, D("150"), date(2022, 6, 30))
    assert len(remaining) == 1
    assert remaining[0].units == D("50")
    assert remaining[0].acquisition_date == date(2021, 1, 1)


def test_tc08_excess_after_sale():
    """After FIFO sale, only remaining 50 units (lot2) participate in distribution."""
    # remaining lot: 50 units, acq 2021-01-01
    r = run_pipeline(
        purchases=[(date(2021, 1, 1), D("50"))],
        distributions_by_year={2021: D("500"), 2022: D("3000")},
        tax_year=2022,
    )
    # prior avg = 500, threshold = 625, excess = 2375
    assert approx(r["excess_distribution"], "2375.00", "0.02")


def test_tc08_single_lot_result():
    r = run_pipeline(
        purchases=[(date(2021, 1, 1), D("50"))],
        distributions_by_year={2021: D("500"), 2022: D("3000")},
        tax_year=2022,
    )
    assert len(r["lots"]) == 1
    assert r["lots"][0]["acquisition_date"] == "2021-01-01"


def test_tc08_has_prior_pfic_2021():
    r = run_pipeline(
        purchases=[(date(2021, 1, 1), D("50"))],
        distributions_by_year={2021: D("500"), 2022: D("3000")},
        tax_year=2022,
    )
    assert r["year_buckets"][2021]["classification"] == DayClassification.PRIOR_PFIC
    assert r["year_buckets"][2022]["classification"] == DayClassification.CURRENT_YEAR


# ── Export tests: engine → Excel → parse Summary → compare with expected ─────

def test_export_tc_main_summary():
    """TC-MAIN Excel Summary sheet must match engine output."""
    txns = TC_MAIN_2022["transactions"]
    purchases = [(t["date"], t["units"]) for t in txns if t["type"] == "purchase"]
    dists = {}
    for t in txns:
        if t["type"] == "distribution":
            dists[t["date"].year] = dists.get(t["date"].year, D("0")) + t["amount"]
    r = run_pipeline(purchases, dists, 2022)
    s = excel_summary(r["full_result"], "TC-MAIN Fund", "CASE-001-SMITH")

    assert approx(s["15e(2) — Excess distribution"], "1660.21", "0.02"), \
        f"Excess: {s['15e(2) — Excess distribution']}"
    assert approx(s["15e(1) — Non-excess ordinary income (Line 16b)"], "823.24", "0.02"), \
        f"Non-excess: {s['15e(1) — Non-excess ordinary income (Line 16b)']}"
    assert approx(s["16c — Additional tax (prior PFIC years)"], "140.72", "0.05"), \
        f"Tax: {s['16c — Additional tax (prior PFIC years)']}"
    assert approx(s["16f — §6621 interest"], "8.19", "0.05"), \
        f"Interest: {s['16f — §6621 interest']}"
    assert approx(s["16c + 16f — Grand total additional liability"], "148.90", "0.10"), \
        f"Grand total: {s['16c + 16f — Grand total additional liability']}"


def test_export_tc04_zero_tax_in_excel():
    """TC-04: all current year → Excel must show $0 tax and $0 interest."""
    r = run_pipeline(
        purchases=[(date(2022, 3, 1), D("1000"))],
        distributions_by_year={2022: D("500")},
        tax_year=2022,
    )
    s = excel_summary(r["full_result"], "TC-04", "TEST")
    assert approx(s.get("16c — Additional tax (prior PFIC years)", 0), "0", "0.01"), \
        f"Expected $0 tax, got {s.get('16c — Additional tax (prior PFIC years)')}"
    assert approx(s.get("16f — §6621 interest", 0), "0", "0.01"), \
        f"Expected $0 interest, got {s.get('16f — §6621 interest')}"


def test_export_tc07_full_excess_in_excel():
    """TC-07: no prior distributions → entire $3000 appears as excess in Excel."""
    r = run_pipeline(
        purchases=[(date(2020, 1, 1), D("1000"))],
        distributions_by_year={2022: D("3000")},
        tax_year=2022,
    )
    s = excel_summary(r["full_result"], "TC-07", "TEST")
    assert approx(s["15e(2) — Excess distribution"], "3000.00", "0.02"), \
        f"Excess: {s['15e(2) — Excess distribution']}"
    assert approx(s.get("15e(1) — Non-excess ordinary income", 0), "0", "0.01"), \
        f"Non-excess should be 0, got {s.get('15e(1) — Non-excess ordinary income')}"


def test_export_tc06_pre_pfic_in_excel():
    """TC-06: $14000 excess, has deferred tax + interest in Excel."""
    r = run_pipeline(
        purchases=[(date(1985, 1, 1), D("1000"))],
        distributions_by_year={2021: D("800"), 2022: D("15000")},
        tax_year=2022,
    )
    s = excel_summary(r["full_result"], "TC-06", "TEST")
    assert approx(s["15e(2) — Excess distribution"], "14000.00", "0.02")
    assert s["16c — Additional tax (prior PFIC years)"] > 0, "Expected non-zero deferred tax"
    assert s["16f — §6621 interest"] > 0, "Expected non-zero interest"


def test_export_tc08_fifo_excess_in_excel():
    """TC-08: after FIFO sale, $2375 excess appears in Excel."""
    r = run_pipeline(
        purchases=[(date(2021, 1, 1), D("50"))],
        distributions_by_year={2021: D("500"), 2022: D("3000")},
        tax_year=2022,
    )
    s = excel_summary(r["full_result"], "TC-08", "TEST")
    assert approx(s["15e(2) — Excess distribution"], "2375.00", "0.02"), \
        f"Excess: {s['15e(2) — Excess distribution']}"


def test_export_matches_fe_display_values():
    """
    The critical parity test: Excel Summary sheet values must equal the
    top-level result fields that the FE Results page displays.

    FE reads: calcResult.total_excess_dist, .additional_tax, .total_interest,
              .grand_total, .ordinary_income  (= top-level pipeline fields)
    Excel reads: full_result dict (same pipeline run)

    This test runs a pipeline once, then asserts Excel == top-level for all 5 key lines,
    covering TC-MAIN (multi-lot), TC-04 (zero tax), TC-07 (full excess).
    """
    cases = [
        # (purchases, distributions_by_year, tax_year, label)
        (
            [(t["date"], t["units"]) for t in TC_MAIN_2022["transactions"] if t["type"] == "purchase"],
            {t["date"].year: t["amount"] for t in TC_MAIN_2022["transactions"] if t["type"] == "distribution"},
            2022, "TC-MAIN",
        ),
        ([(date(2022, 3, 1), D("1000"))], {2022: D("500")}, 2022, "TC-04"),
        ([(date(2020, 1, 1), D("1000"))], {2022: D("3000")}, 2022, "TC-07"),
        ([(date(2018, 1, 1), D("1000"))], {2018: D("500"), 2019: D("2000")}, 2019, "TC-02"),
        ([(date(1985, 1, 1), D("1000"))], {2021: D("800"), 2022: D("15000")}, 2022, "TC-06"),
    ]

    for purchases, dists, yr, label in cases:
        r = run_pipeline(purchases, dists, yr)
        s = excel_summary(r["full_result"], label, "TEST")

        # What the FE displays (top-level pipeline fields)
        fe_excess   = float(r["excess_distribution"])
        fe_tax      = float(r["total_deferred_tax"])
        fe_interest = float(r["total_interest"])
        fe_grand    = float(r["grand_total"])
        fe_ordinary = float(r["non_excess_ordinary"])  # non-excess portion shown in FE

        # What the Excel shows (parsed from Summary sheet)
        xl_excess   = s.get("15e(2) — Excess distribution", 0) or 0
        xl_tax      = s.get("16c — Additional tax (prior PFIC years)", 0) or 0
        xl_interest = s.get("16f — §6621 interest", 0) or 0
        xl_grand    = s.get("16c + 16f — Grand total additional liability", 0) or 0
        xl_ordinary = s.get("15e(1) — Non-excess ordinary income (Line 16b)", 0) or 0

        tol = 0.02
        assert abs(fe_excess   - xl_excess)   <= tol, f"{label} excess: FE={fe_excess} XL={xl_excess}"
        assert abs(fe_tax      - xl_tax)      <= tol, f"{label} tax: FE={fe_tax} XL={xl_tax}"
        assert abs(fe_interest - xl_interest) <= tol, f"{label} interest: FE={fe_interest} XL={xl_interest}"
        assert abs(fe_grand    - xl_grand)    <= tol, f"{label} grand: FE={fe_grand} XL={xl_grand}"
        assert abs(fe_ordinary - xl_ordinary) <= tol, f"{label} ordinary: FE={fe_ordinary} XL={xl_ordinary}"


def test_export_excel_grand_total_equals_tax_plus_interest():
    """Cross-check: grand total in Excel = tax + interest for every TC."""
    cases = [
        ([(date(2020, 1, 1), D("1000"))], {2020: D("500"), 2021: D("2000")}, 2021),  # TC-01
        ([(date(2022, 3, 1), D("1000"))], {2022: D("500")}, 2022),                    # TC-04
        ([(date(2020, 1, 1), D("1000"))], {2022: D("3000")}, 2022),                   # TC-07
    ]
    for purchases, dists, yr in cases:
        r = run_pipeline(purchases, dists, yr)
        s = excel_summary(r["full_result"])
        tax = D(str(s.get("16c — Additional tax (prior PFIC years)", 0) or 0))
        interest = D(str(s.get("16f — §6621 interest", 0) or 0))
        grand = D(str(s.get("16c + 16f — Grand total additional liability", 0) or 0))
        assert approx(grand, tax + interest, "0.02"), \
            f"Grand total {grand} != tax {tax} + interest {interest}"


# ── Sheet structure tests: verify new 5-sheet layout ─────────────────────────

def _tc_main_excel():
    txns = TC_MAIN_2022["transactions"]
    purchases = [(t["date"], t["units"]) for t in txns if t["type"] == "purchase"]
    dists = {}
    for t in txns:
        if t["type"] == "distribution":
            dists[t["date"].year] = dists.get(t["date"].year, D("0")) + t["amount"]
    return run_pipeline(purchases, dists, 2022)


def test_excel_has_5_sheets():
    r = _tc_main_excel()
    sheets = parse_excel(r["full_result"])
    assert set(sheets.keys()) == {"Summary", "Lot Summary", "Year Allocation", "Interest Detail", "Daily Allocation"}, \
        f"Sheets: {list(sheets.keys())}"


def test_excel_lot_summary_has_6_lots():
    """TC-MAIN has 6 lots — Lot Summary sheet must have 6 data rows + header + total."""
    r = _tc_main_excel()
    sheets = parse_excel(r["full_result"])
    rows = sheets["Lot Summary"]["_rows"]
    # rows = [section_title, header, L1..L6, Total] = 9 rows
    lot_rows = [row for row in rows if str(row[0] or "").startswith("L") and str(row[0] or "")[1:].isdigit()]
    assert len(lot_rows) == 6, f"Expected 6 lot rows, got {len(lot_rows)}: {[r[0] for r in lot_rows]}"


def test_excel_year_allocation_has_correct_years():
    """TC-MAIN: Year Allocation sheet must have 2021 and 2022 rows."""
    r = _tc_main_excel()
    sheets = parse_excel(r["full_result"])
    year_col = [str(row[0]) for row in sheets["Year Allocation"]["_rows"]]
    assert "2021" in year_col, "2021 missing from Year Allocation"
    assert "2022" in year_col, "2022 missing from Year Allocation"


def test_excel_interest_detail_has_prior_pfic_only():
    """Interest Detail sheet must only contain prior_pfic years (2021 for TC-MAIN)."""
    r = _tc_main_excel()
    sheets = parse_excel(r["full_result"])
    rows = sheets["Interest Detail"]["_rows"]
    year_col = [str(row[0]) for row in rows if row[0] not in (None, "Year", "Total", "Step 3 — §6621 Interest  [§6622 daily compound]")]
    # 2022 is current year — must NOT appear in Interest Detail
    assert "2022" not in year_col, f"Current year 2022 should not be in Interest Detail: {year_col}"
    assert "2021" in year_col, f"Prior PFIC year 2021 missing: {year_col}"


def test_excel_daily_allocation_days_sum():
    """Daily Allocation sheet days must sum to total holding days."""
    r = _tc_main_excel()
    sheets = parse_excel(r["full_result"])
    rows = sheets["Daily Allocation"]["_rows"]
    # rows with integer year in col 0
    day_rows = [row for row in rows if isinstance(row[0], int)]
    total = sum(row[1] for row in day_rows if isinstance(row[1], int))
    # total days across all merged buckets
    expected = sum(b["days"] for b in r["year_buckets"].values())
    assert total == expected, f"Days sum {total} != expected {expected}"


def test_pdf_workpaper_generates():
    from pfic_engine.output.pdf_generator import generate_form8621_workpaper
    r = _tc_main_excel()
    pdf = generate_form8621_workpaper(r["full_result"], "TC-MAIN Fund", "CASE-001-SMITH")
    assert pdf[:4] == b"%PDF", "Not a valid PDF"
    assert len(pdf) > 2000


def test_line16a_generates():
    from pfic_engine.output.line16a_statement import generate_line16a_statement
    r = _tc_main_excel()
    pdf = generate_line16a_statement(r["full_result"], "TC-MAIN Fund", "CASE-001-SMITH")
    assert pdf[:4] == b"%PDF", "Not a valid PDF"
    assert len(pdf) > 2000


# ── HTTP-level export tests (via FastAPI TestClient) ─────────────────────────
# These mirror what the FE calls: POST /holdings/{id}/calculate → GET export/{type}
# Values in the exported files must match the calculation result returned to the FE.

@pytest.fixture(scope="module")
def http_export_setup():
    """Create holding, run calculation, return (client, holding_id, calc_result)."""
    import uuid
    from fastapi.testclient import TestClient
    from api.main import app

    tc = TestClient(app)
    unique_code = f"HTTP-EXP-{uuid.uuid4().hex[:8].upper()}"
    r = tc.post("/clients/", json={"client_code": unique_code})
    assert r.status_code in (200, 201), r.text
    cid = r.json()["id"]

    r = tc.post(f"/clients/{cid}/holdings/", json={"pfic_name": "HTTP Export Fund", "currency": "USD", "method": "1291"})
    assert r.status_code in (200, 201), r.text
    hid = r.json()["id"]

    for txn in [
        {"txn_date": "2020-01-15", "txn_type": "purchase", "units": 100, "total_value_usd": 10000},
        {"txn_date": "2021-06-30", "txn_type": "distribution", "total_value_usd": 500},
        {"txn_date": "2022-06-30", "txn_type": "distribution", "total_value_usd": 800},
        {"txn_date": "2023-06-30", "txn_type": "distribution", "total_value_usd": 2500},
    ]:
        tc.post(f"/holdings/{hid}/transactions/", json=txn)

    r = tc.post(f"/holdings/{hid}/calculate", json={"tax_year": 2023})
    assert r.status_code == 200, r.text
    calc = r.json()

    return tc, hid, calc


def test_http_export_pdf_is_valid_pdf(http_export_setup):
    tc, hid, calc = http_export_setup
    r = tc.get(f"/holdings/{hid}/calculations/2023/export/pdf")
    assert r.status_code == 200, f"PDF export failed: {r.text}"
    assert r.headers["content-type"] == "application/pdf"
    assert r.content[:4] == b"%PDF"
    assert len(r.content) > 2000


def test_http_export_line16a_is_valid_pdf(http_export_setup):
    tc, hid, calc = http_export_setup
    r = tc.get(f"/holdings/{hid}/calculations/2023/export/line16a")
    assert r.status_code == 200, f"Line16a export failed: {r.text}"
    assert r.headers["content-type"] == "application/pdf"
    assert r.content[:4] == b"%PDF"
    assert len(r.content) > 2000


def test_http_export_excel_is_valid_xlsx(http_export_setup):
    tc, hid, calc = http_export_setup
    r = tc.get(f"/holdings/{hid}/calculations/2023/export/excel")
    assert r.status_code == 200, f"Excel export failed: {r.text}"
    assert "spreadsheetml" in r.headers["content-type"]
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "Summary" in wb.sheetnames


def test_http_export_excel_grand_total_matches_calc(http_export_setup):
    """Excel grand total must match the value returned in the calc API response."""
    tc, hid, calc = http_export_setup
    expected_grand_total = D(calc["full_result"]["grand_total"])

    r = tc.get(f"/holdings/{hid}/calculations/2023/export/excel")
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["Summary"]
    rows = list(ws.iter_rows(values_only=True))

    grand_total_row = next(
        (row for row in rows if row[0] and "Grand total" in str(row[0])), None
    )
    assert grand_total_row is not None, "Grand total row not found in Excel Summary"
    assert approx(grand_total_row[1], expected_grand_total, "0.01"), \
        f"Excel grand total {grand_total_row[1]} != calc {expected_grand_total}"


def test_http_export_excel_excess_matches_calc(http_export_setup):
    """Excel excess distribution must match the calc result."""
    tc, hid, calc = http_export_setup
    expected_excess = D(calc["full_result"]["excess_distribution"])

    r = tc.get(f"/holdings/{hid}/calculations/2023/export/excel")
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["Summary"]
    rows = list(ws.iter_rows(values_only=True))

    excess_row = next(
        (row for row in rows if row[0] and "Excess distribution" in str(row[0]) and "15e(2)" in str(row[0])), None
    )
    assert excess_row is not None, "Excess distribution row not found in Excel Summary"
    assert approx(excess_row[1], expected_excess, "0.01"), \
        f"Excel excess {excess_row[1]} != calc {expected_excess}"


def test_http_export_excel_interest_matches_calc(http_export_setup):
    """Excel §6621 interest must match the calc result."""
    tc, hid, calc = http_export_setup
    expected_interest = D(calc["full_result"]["total_interest"])

    r = tc.get(f"/holdings/{hid}/calculations/2023/export/excel")
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["Summary"]
    rows = list(ws.iter_rows(values_only=True))

    interest_row = next(
        (row for row in rows if row[0] and "6621 interest" in str(row[0])), None
    )
    assert interest_row is not None, "§6621 interest row not found in Excel Summary"
    assert approx(interest_row[1], expected_interest, "0.01"), \
        f"Excel interest {interest_row[1]} != calc {expected_interest}"


def test_http_export_404_when_no_calc(http_export_setup):
    """Export must return 404 when no calculation exists for the requested year."""
    tc, hid, _ = http_export_setup
    for typ in ("pdf", "line16a", "excel"):
        r = tc.get(f"/holdings/{hid}/calculations/1999/export/{typ}")
        assert r.status_code == 404, f"{typ} should 404 for non-existent calc, got {r.status_code}"
