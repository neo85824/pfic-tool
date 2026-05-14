"""
Excel calculation workpapers — mirrors the FE Result page structure 1:1.

Sheet 1: Summary          — 125% Test + Form 8621 Part V lines
Sheet 2: Lot Summary      — Excess Distribution by Lot table
Sheet 3: Year Allocation  — Step 2 Year-by-Year Allocation & Tax (lots merged)
Sheet 4: Interest Detail  — Step 3 §6621 Interest by year
Sheet 5: Daily Allocation — Year / Days / Amount per year
"""
from decimal import Decimal
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Palette ──────────────────────────────────────────────────────────────────
DARK_BLUE   = "1A365D"
MID_BLUE    = "2D6A9F"
LIGHT_BLUE  = "EEF4FB"
ORANGE_BG   = "FFF3CD"
WHITE       = "FFFFFF"
GREY        = "F7F7F7"
GREEN_BG    = "F0FFF4"
RED_TEXT    = "C0392B"


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(ws, row, n_cols, color=DARK_BLUE):
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF", size=9)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")
        c.border = _thin_border()


def _section_title(ws, row, text, n_cols):
    ws.cell(row=row, column=1, value=text).font = Font(bold=True, size=10, color=DARK_BLUE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)


def _money(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=float(Decimal(str(value))) if value is not None else 0.0)
    c.number_format = '#,##0.00'
    c.border = _thin_border()
    return c


def _text(ws, row, col, value, bold=False, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=9)
    c.alignment = Alignment(horizontal=align)
    c.border = _thin_border()
    return c


def _row_fill(ws, row, n_cols, color):
    fill = PatternFill("solid", fgColor=color)
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).fill = fill


def _total_row(ws, row, n_cols):
    fill = PatternFill("solid", fgColor="E8EEF4")
    font = Font(bold=True, size=9)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.border = _thin_border()


def _D(v):
    try:
        return Decimal(str(v))
    except Exception:
        return Decimal("0")


# ── Merge helpers ─────────────────────────────────────────────────────────────

def _merged_year_results(deferred_tax_results):
    """Combine year_results across all lots → {year_int: merged_dict}."""
    merged = {}
    for dr in deferred_tax_results:
        for yr_key, yr_data in dr["year_results"].items():
            yr = int(yr_key)
            if yr not in merged:
                merged[yr] = {
                    "classification": yr_data["classification"],
                    "allocated_amount": _D(yr_data["allocated_amount"]),
                    "tax": _D(yr_data.get("tax") or 0),
                    "interest": _D(yr_data.get("interest") or 0),
                    "interest_start": yr_data.get("interest_start", ""),
                    "interest_end": yr_data.get("interest_end", ""),
                }
            else:
                merged[yr]["allocated_amount"] += _D(yr_data["allocated_amount"])
                merged[yr]["tax"]              += _D(yr_data.get("tax") or 0)
                merged[yr]["interest"]         += _D(yr_data.get("interest") or 0)
    return merged


# ── Sheet 1: Summary ──────────────────────────────────────────────────────────

def _sheet_summary(wb, full_result, holding_name, client_code):
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18

    tax_year = full_result.get("tax_year", "")
    prior_avg = _D(full_result.get("prior_3yr_average", "0"))
    cur_dist  = _D(full_result.get("current_year_distribution", "0"))
    non_excess = _D(full_result.get("non_excess_ordinary", "0"))
    excess    = _D(full_result.get("excess_distribution", "0"))
    total_tax = _D(full_result.get("total_deferred_tax", "0"))
    total_int = _D(full_result.get("total_interest", "0"))
    grand     = _D(full_result.get("grand_total", "0"))
    ordinary  = _D(full_result.get("total_ordinary_income", "0"))

    r = 1
    ws.cell(r, 1, "PFIC Form 8621 — Calculation Workpaper").font = Font(bold=True, size=13, color=DARK_BLUE)
    ws.merge_cells(f"A{r}:B{r}")
    r += 1
    ws.cell(r, 1, f"Tax Year: {tax_year}  |  PFIC: {holding_name}  |  Client: {client_code}").font = Font(size=9, color="444444")
    ws.merge_cells(f"A{r}:B{r}")
    r += 1
    ws.cell(r, 1, "Method: §1291 Excess Distribution  |  Currency: USD").font = Font(size=9, color="444444")
    ws.merge_cells(f"A{r}:B{r}")
    r += 2

    # ── 125% Test ──
    _section_title(ws, r, "125% Test  [IRC §1291(b)(2)(A)]", 2); r += 1
    _hdr(ws, r, 2); ws.cell(r, 1, "Item").font = Font(bold=True, color="FFFFFF", size=9)
    ws.cell(r, 2, "Amount (USD)").font = Font(bold=True, color="FFFFFF", size=9)
    ws.cell(r, 2).alignment = Alignment(horizontal="right"); r += 1

    test_rows = [
        ("Current year distribution (15c)", cur_dist, WHITE),
        ("Prior 3-year average (15b)", prior_avg, WHITE),
        ("Benchmark — 125% threshold", prior_avg * Decimal("1.25"), GREY),
        ("Excess distribution (15e(2))", excess, ORANGE_BG),
        ("Non-excess ordinary income (15e(1))", non_excess, WHITE),
    ]
    for label, val, bg in test_rows:
        _text(ws, r, 1, label)
        _money(ws, r, 2, val).alignment = Alignment(horizontal="right")
        _row_fill(ws, r, 2, bg)
        r += 1

    r += 1

    # ── Form 8621 Part V Summary ──
    _section_title(ws, r, "Form 8621 Part V — Summary", 2); r += 1
    _hdr(ws, r, 2)
    ws.cell(r, 1, "Line / Description").font = Font(bold=True, color="FFFFFF", size=9)
    ws.cell(r, 2, "Amount (USD)").font = Font(bold=True, color="FFFFFF", size=9)
    ws.cell(r, 2).alignment = Alignment(horizontal="right"); r += 1

    pv_rows = [
        ("15e(1) — Non-excess ordinary income (Line 16b)", non_excess, WHITE),
        ("15e(2) — Excess distribution", excess, WHITE),
        ("16c — Additional tax (prior PFIC years)", total_tax, WHITE),
        ("16f — §6621 interest", total_int, WHITE),
        ("16c + 16f — Grand total additional liability", grand, ORANGE_BG),
        ("16b — Total ordinary income (pre-PFIC + current year)", ordinary, GREY),
    ]
    for label, val, bg in pv_rows:
        bold = bg == ORANGE_BG
        c1 = _text(ws, r, 1, label, bold=bold)
        c2 = _money(ws, r, 2, val)
        c2.alignment = Alignment(horizontal="right")
        if bold:
            c1.font = Font(bold=True, size=9)
            c2.font = Font(bold=True, size=9)
        _row_fill(ws, r, 2, bg)
        r += 1


# ── Sheet 2: Lot Summary ──────────────────────────────────────────────────────

def _sheet_lot_summary(wb, full_result):
    ws = wb.create_sheet("Lot Summary")
    cols = {"A": 6, "B": 14, "C": 14, "D": 16, "E": 16, "F": 16, "G": 16}
    for col, w in cols.items():
        ws.column_dimensions[col].width = w

    deferred = full_result.get("deferred_tax_results", [])

    r = 1
    _section_title(ws, r, "Excess Distribution by Lot  [§1291(a)(1) proportional allocation]", 7); r += 1

    headers = ["Lot", "Acquired", "Units", "Excess Share", "Deferred Tax (16c)", "§6621 Interest (16f)", "Grand Total"]
    _hdr(ws, r, 7)
    for i, h in enumerate(headers, 1):
        ws.cell(r, i, h).font = Font(bold=True, color="FFFFFF", size=9)
        ws.cell(r, i).alignment = Alignment(horizontal="right" if i > 2 else "center")
    r += 1

    tot_excess = tot_tax = tot_int = tot_grand = Decimal("0")

    for i, dr in enumerate(deferred, 1):
        bg = WHITE if i % 2 == 0 else LIGHT_BLUE
        lot_excess = _D(dr.get("lot_excess", 0))
        tax        = _D(dr.get("total_deferred_tax", 0))
        interest   = _D(dr.get("total_interest", 0))
        grand      = _D(dr.get("grand_total", 0))
        units      = _D(dr.get("units", 0))

        _text(ws, r, 1, f"L{i}", align="center")
        _text(ws, r, 2, dr.get("acquisition_date", ""), align="center")
        _money(ws, r, 3, units).alignment = Alignment(horizontal="right")
        _money(ws, r, 4, lot_excess).alignment = Alignment(horizontal="right")
        _money(ws, r, 5, tax).alignment = Alignment(horizontal="right")
        _money(ws, r, 6, interest).alignment = Alignment(horizontal="right")
        _money(ws, r, 7, grand).alignment = Alignment(horizontal="right")
        _row_fill(ws, r, 7, bg)

        tot_excess += lot_excess; tot_tax += tax; tot_int += interest; tot_grand += grand
        r += 1

    _total_row(ws, r, 7)
    ws.cell(r, 1, "Total").font = Font(bold=True, size=9)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    for col, val in [(4, tot_excess), (5, tot_tax), (6, tot_int), (7, tot_grand)]:
        c = _money(ws, r, col, val)
        c.font = Font(bold=True, size=9)
        c.alignment = Alignment(horizontal="right")


# ── Sheet 3: Year Allocation ──────────────────────────────────────────────────

def _sheet_year_allocation(wb, full_result):
    ws = wb.create_sheet("Year Allocation")
    for col, w in zip("ABCDE", [8, 8, 16, 10, 16]):
        ws.column_dimensions[col].width = w

    year_buckets    = full_result.get("year_buckets", {})
    deferred        = full_result.get("deferred_tax_results", [])
    merged          = _merged_year_results(deferred)

    r = 1
    _section_title(ws, r, "Step 2 — Year-by-Year Allocation & Tax  [§1291(c)(2)]", 5); r += 1

    headers = ["Year", "Days", "Allocated Amount", "Rate", "Deferred Tax"]
    _hdr(ws, r, 5)
    for i, h in enumerate(headers, 1):
        ws.cell(r, i, h).font = Font(bold=True, color="FFFFFF", size=9)
        ws.cell(r, i).alignment = Alignment(horizontal="right" if i > 1 else "center")
    r += 1

    sorted_years = sorted(int(y) for y in year_buckets)
    tot_days = tot_alloc = tot_tax = Decimal("0")

    for i, yr in enumerate(sorted_years):
        bucket  = year_buckets[str(yr)]
        days    = bucket["days"]
        amount  = _D(bucket["amount"])
        cls     = bucket["classification"]
        is_cur  = cls == "current_year"
        is_pre  = cls == "pre_pfic"

        yr_data = merged.get(yr, {})
        tax     = yr_data.get("tax", Decimal("0")) if not is_cur and not is_pre else Decimal("0")
        rate    = (tax / amount * 100) if amount > 0 and not is_cur and not is_pre else None

        bg = GREY if is_cur else (WHITE if i % 2 == 0 else LIGHT_BLUE)

        _text(ws, r, 1, str(yr), align="center")
        ws.cell(r, 2, days).border = _thin_border()
        ws.cell(r, 2).alignment = Alignment(horizontal="right")
        _money(ws, r, 3, amount).alignment = Alignment(horizontal="right")

        if rate is not None:
            ws.cell(r, 4, float(rate) / 100).number_format = '0.0%'
        else:
            ws.cell(r, 4, "ordinary income" if is_cur else ("pre-PFIC" if is_pre else "—"))
        ws.cell(r, 4).border = _thin_border()
        ws.cell(r, 4).alignment = Alignment(horizontal="right")

        if is_cur or is_pre:
            ws.cell(r, 5, "—").border = _thin_border()
            ws.cell(r, 5).alignment = Alignment(horizontal="right")
        else:
            _money(ws, r, 5, tax).alignment = Alignment(horizontal="right")

        _row_fill(ws, r, 5, bg)
        tot_days += days; tot_alloc += amount; tot_tax += tax
        r += 1

    _total_row(ws, r, 5)
    ws.cell(r, 1, "Total").font = Font(bold=True, size=9)
    ws.cell(r, 2, int(tot_days)).font = Font(bold=True, size=9)
    ws.cell(r, 2).alignment = Alignment(horizontal="right")
    c3 = _money(ws, r, 3, tot_alloc); c3.font = Font(bold=True, size=9); c3.alignment = Alignment(horizontal="right")
    ws.cell(r, 4, "").border = _thin_border()
    c5 = _money(ws, r, 5, tot_tax);  c5.font = Font(bold=True, size=9); c5.alignment = Alignment(horizontal="right")


# ── Sheet 4: Interest Detail ──────────────────────────────────────────────────

def _sheet_interest_detail(wb, full_result):
    ws = wb.create_sheet("Interest Detail")
    for col, w in zip("ABCD", [8, 16, 18, 16]):
        ws.column_dimensions[col].width = w

    deferred  = full_result.get("deferred_tax_results", [])
    merged    = _merged_year_results(deferred)
    int_end   = None

    r = 1
    # Find interest_end (same for all years — current return due date)
    for yr_data in merged.values():
        if yr_data.get("interest_end"):
            int_end = yr_data["interest_end"][:10]
            break

    title = f"Step 3 — §6621 Interest  [§6622 daily compound]"
    if int_end:
        title += f"  |  Interest end: {int_end}"
    _section_title(ws, r, title, 4); r += 1

    headers = ["Year", "Deferred Tax", "Filing Deadline", "§6621 Interest"]
    _hdr(ws, r, 4)
    for i, h in enumerate(headers, 1):
        ws.cell(r, i, h).font = Font(bold=True, color="FFFFFF", size=9)
        ws.cell(r, i).alignment = Alignment(horizontal="right" if i != 3 else "center")
    r += 1

    prior_years = sorted(yr for yr, d in merged.items() if d["classification"] == "prior_pfic")
    tot_tax = tot_int = Decimal("0")

    for i, yr in enumerate(prior_years):
        d   = merged[yr]
        tax = d["tax"]
        interest = d["interest"]
        i_start  = d.get("interest_start", "")[:10]
        is_covid = i_start and i_start[5:7] not in ("04",)

        bg = WHITE if i % 2 == 0 else LIGHT_BLUE

        _text(ws, r, 1, str(yr), align="center")
        _money(ws, r, 2, tax).alignment = Alignment(horizontal="right")
        label = i_start + (" ⚠ COVID" if is_covid else "")
        _text(ws, r, 3, label, align="center")
        _money(ws, r, 4, interest).alignment = Alignment(horizontal="right")
        _row_fill(ws, r, 4, bg)

        tot_tax += tax; tot_int += interest
        r += 1

    _total_row(ws, r, 4)
    ws.cell(r, 1, "Total").font = Font(bold=True, size=9)
    c2 = _money(ws, r, 2, tot_tax);  c2.font = Font(bold=True, size=9); c2.alignment = Alignment(horizontal="right")
    ws.cell(r, 3, "").border = _thin_border()
    c4 = _money(ws, r, 4, tot_int); c4.font = Font(bold=True, size=9); c4.alignment = Alignment(horizontal="right")


# ── Sheet 5: Daily Allocation ─────────────────────────────────────────────────

def _sheet_daily_allocation(wb, full_result):
    ws = wb.create_sheet("Daily Allocation")
    for col, w in zip("ABCD", [10, 8, 20, 18]):
        ws.column_dimensions[col].width = w

    year_buckets = full_result.get("year_buckets", {})

    r = 1
    _section_title(ws, r, "Daily Allocation Summary — Year Buckets  [§1291(a)(1)(A)]", 4); r += 1

    headers = ["Year", "Days", "Amount Allocated", "Classification"]
    _hdr(ws, r, 4)
    for i, h in enumerate(headers, 1):
        ws.cell(r, i, h).font = Font(bold=True, color="FFFFFF", size=9)
        ws.cell(r, i).alignment = Alignment(horizontal="right" if i == 2 else "center")
    r += 1

    for i, (yr_str, bucket) in enumerate(sorted(year_buckets.items(), key=lambda x: int(x[0])), 1):
        bg = WHITE if i % 2 == 0 else GREY
        _text(ws, r, 1, int(yr_str), align="center")
        ws.cell(r, 2, bucket["days"]).border = _thin_border()
        ws.cell(r, 2).alignment = Alignment(horizontal="right")
        _money(ws, r, 3, bucket["amount"]).alignment = Alignment(horizontal="right")
        _text(ws, r, 4, bucket["classification"].replace("_", " "), align="center")
        _row_fill(ws, r, 4, bg)
        r += 1


# ── Public API ────────────────────────────────────────────────────────────────

def generate_workpapers(full_result: dict, holding_name: str, client_code: str) -> bytes:
    wb = openpyxl.Workbook()
    _sheet_summary(wb, full_result, holding_name, client_code)
    _sheet_lot_summary(wb, full_result)
    _sheet_year_allocation(wb, full_result)
    _sheet_interest_detail(wb, full_result)
    _sheet_daily_allocation(wb, full_result)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
